"""
Point d'entrée du sidecar Python pour TELETAXI CRM.

Protocole : JSON line-delimited sur stdin/stdout.
Chaque ligne stdin = une requête JSON-RPC.
Chaque ligne stdout = une réponse ou événement JSON.
Stderr = logs internes (non parsés par Tauri).
"""

from __future__ import annotations

import json
import logging
import sys
import traceback
from pathlib import Path
from typing import Any


# ─── Résolution des chemins ──────────────────────────────────────────────────
#
# Deux contextes :
#   1. Mode PyInstaller (frozen) : les modules teletaxi_import sont gelés dans
#      le bundle, et les fichiers statiques (JARs) sont dans sys._MEIPASS/lib/
#   2. Mode dev (source) : teletaxi_import est dans le repo adjacent

_IS_FROZEN = getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS")
_HERE = Path(__file__).parent.resolve()
_TELETAXI_REPO = _HERE.parent.parent / "teletaxi_import"  # TeleTaxi/teletaxi_import/

# En mode source : ajouter le repo teletaxi_import au sys.path
# En mode PyInstaller : les modules sont gelés dans le bundle
if _IS_FROZEN:
    sys.path.insert(0, sys._MEIPASS)  # type: ignore[attr-defined]
else:
    sys.path.insert(0, str(_TELETAXI_REPO))


def _get_lib_dir() -> str:
    """Retourne le chemin vers le dossier contenant les JARs UCanAccess."""
    if _IS_FROZEN:
        return str(Path(sys._MEIPASS) / "lib")  # type: ignore[attr-defined]
    return str(_TELETAXI_REPO / "lib")


def _get_log_dir() -> Path:
    """Retourne le dossier de logs de l'application."""
    if sys.platform == "darwin":
        base = Path.home() / "Library" / "Logs" / "TeleTaxiCRM"
    elif sys.platform == "win32":
        base = Path.home() / "AppData" / "Local" / "TeleTaxiCRM" / "logs"
    else:
        base = Path.home() / ".local" / "share" / "TeleTaxiCRM" / "logs"
    base.mkdir(parents=True, exist_ok=True)
    return base


# Configurer le logging AVANT tout import teletaxi_import.
# Le logger "teletaxi" utilise sys.stdout par défaut (voir logger.py).
# On force stderr ici pour ne pas polluer le canal JSON stdout.
logging.basicConfig(stream=sys.stderr, level=logging.WARNING)


def _patch_teletaxi_logger_to_stderr() -> None:
    """Redirige les StreamHandlers stdout du logger 'teletaxi' vers stderr."""
    logger = logging.getLogger("teletaxi")
    for h in logger.handlers:
        if isinstance(h, logging.StreamHandler) and h.stream is sys.stdout:
            h.stream = sys.stderr


# ─── Helpers I/O ─────────────────────────────────────────────────────────────


def write_message(msg: dict[str, Any]) -> None:
    """Émet un message JSON sur stdout (line-delimited)."""
    sys.stdout.write(json.dumps(msg, ensure_ascii=False, default=str) + "\n")
    sys.stdout.flush()


def write_error(
    request_id: str, code: str, message: str, details: str = ""
) -> None:
    write_message(
        {
            "id": request_id,
            "type": "error",
            "code": code,
            "message": message,
            "details": details,
        }
    )


def write_progress(
    request_id: str,
    phase: str,
    current: int = 0,
    total: int = 0,
    message: str = "",
) -> None:
    write_message(
        {
            "id": request_id,
            "type": "progress",
            "phase": phase,
            "current": current,
            "total": total,
            "message": message,
        }
    )


# ─── Handlers JSON-RPC ───────────────────────────────────────────────────────


def handle_ping(request_id: str, params: dict) -> None:
    """Sanity check : vérifie que le sidecar répond."""
    write_message(
        {
            "id": request_id,
            "type": "result",
            "data": {
                "status": "ok",
                "version": "0.1.0",
                "python": sys.version,
                "lib_dir": _get_lib_dir(),
            },
        }
    )


def handle_test_connection(request_id: str, params: dict) -> None:
    """Teste la connexion à une base Access via UCanAccess/JDBC."""
    accdb_path = params.get("accdb_path")
    if not accdb_path:
        write_error(request_id, "MISSING_PARAM", "accdb_path requis")
        return

    try:
        from teletaxi_import.config import AppConfig
        from teletaxi_import.database.connection import AccessConnection

        config = AppConfig.from_env(
            accdb_path=accdb_path,
            excel_path="/tmp/dummy_sidecar",
            lib_dir=_get_lib_dir(),
        )

        with AccessConnection(config, read_only=True) as conn:
            tables = conn.list_tables()
            counts: dict[str, int] = {}
            for table in ["PRESCRIPTEUR", "PATIENT", "TRANSPORT"]:
                if table in tables:
                    result = conn.fetch_one(f"SELECT COUNT(*) AS nb FROM [{table}]")
                    if result:
                        nb = (
                            result.get("nb")
                            or result.get("NB")
                            or next(iter(result.values()), 0)
                        )
                        counts[table] = int(nb or 0)

        write_message(
            {
                "id": request_id,
                "type": "result",
                "data": {
                    "connected": True,
                    "tables_count": len(tables),
                    "counts": counts,
                },
            }
        )
    except Exception as e:
        write_error(
            request_id, "CONNECTION_FAILED", str(e), traceback.format_exc()
        )


def handle_preview_excel(request_id: str, params: dict) -> None:
    """Lit les premières lignes d'une feuille Excel sans lancer l'import."""
    excel_path = params.get("excel_path")
    sheet_name = params.get("sheet_name", "Prescripteurs")
    limit = int(params.get("limit", 50))

    if not excel_path:
        write_error(request_id, "MISSING_PARAM", "excel_path requis")
        return

    try:
        import openpyxl
        from teletaxi_import.config import DATA_START_ROW
        from teletaxi_import.excel.sheets import (
            BENEFICIAIRES_SCHEMA,
            COURSES_SCHEMA,
            PRESCRIPTEURS_SCHEMA,
        )

        # Mapping nom de feuille → colonnes nommées
        SCHEMAS = {
            s.name: s.column_mapping
            for s in [PRESCRIPTEURS_SCHEMA, BENEFICIAIRES_SCHEMA, COURSES_SCHEMA]
        }
        col_map = SCHEMAS.get(sheet_name)

        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            write_error(
                request_id,
                "SHEET_NOT_FOUND",
                f"Feuille '{sheet_name}' introuvable dans ce fichier Excel",
            )
            return

        ws = wb[sheet_name]
        raw_rows = list(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True)
        )
        wb.close()

        # Filtrer les lignes entièrement vides
        indexed = [
            (i, row)
            for i, row in enumerate(raw_rows)
            if any(v is not None for v in row)
        ]
        total_rows = len(indexed)

        preview_rows = []
        for i, (offset, row) in enumerate(indexed[:limit]):
            row_num = DATA_START_ROW + offset
            if col_map:
                data = {
                    col_map[j]: (str(v) if v is not None else None)
                    for j, v in enumerate(row)
                    if j in col_map
                }
            else:
                data = {
                    str(j): (str(v) if v is not None else None)
                    for j, v in enumerate(row)
                }
            preview_rows.append(
                {"row_number": row_num, "status": "valid", "data": data}
            )

        write_message(
            {
                "id": request_id,
                "type": "result",
                "data": {
                    "sheet_name": sheet_name,
                    "total_rows": total_rows,
                    "preview_rows": preview_rows,
                },
            }
        )
    except Exception as e:
        write_error(request_id, "PREVIEW_FAILED", str(e), traceback.format_exc())


def handle_run_import(request_id: str, params: dict) -> None:
    """
    Lance un import complet Excel → Access.

    NOTE Phase 2 : pas de streaming de progression (callbacks = Phase 3).
    """
    excel_path = params.get("excel_path")
    accdb_path = params.get("accdb_path")
    backup_enabled = bool(params.get("backup_enabled", True))

    if not excel_path or not accdb_path:
        write_error(request_id, "MISSING_PARAM", "excel_path et accdb_path requis")
        return

    try:
        from teletaxi_import.config import AppConfig
        from teletaxi_import.logger import setup_logger
        from teletaxi_import.services.import_service import ImportService

        config = AppConfig.from_env(
            excel_path=excel_path,
            accdb_path=accdb_path,
            backup_enabled=backup_enabled,
            lib_dir=_get_lib_dir(),
            log_dir=str(_get_log_dir()),
        )
        setup_logger(config.log_dir)
        # Patch immédiat : le logger "teletaxi" ne doit pas écrire sur stdout
        _patch_teletaxi_logger_to_stderr()

        write_progress(request_id, "starting", message="Démarrage de l'import")

        # TODO Phase 3 : adapter ImportService avec callbacks on_progress
        service = ImportService(config)
        report = service.run()

        write_message(
            {
                "id": request_id,
                "type": "result",
                "data": {
                    "report": {
                        "prescripteurs": {
                            "inseres": report.prescripteurs.inseres,
                            "ignores": report.prescripteurs.ignores,
                            "erreurs": report.prescripteurs.erreurs,
                        },
                        "beneficiaires": {
                            "inseres": report.beneficiaires.inseres,
                            "ignores": report.beneficiaires.ignores,
                            "erreurs": report.beneficiaires.erreurs,
                        },
                        "courses": {
                            "inseres": report.courses.inseres,
                            "ignores": report.courses.ignores,
                            "erreurs": report.courses.erreurs,
                        },
                        "total_inseres": report.total_inseres,
                        "total_erreurs": report.total_erreurs,
                    }
                },
            }
        )
    except Exception as e:
        write_error(request_id, "IMPORT_FAILED", str(e), traceback.format_exc())


# ─── Dispatch ────────────────────────────────────────────────────────────────

HANDLERS: dict[str, Any] = {
    "ping": handle_ping,
    "test_connection": handle_test_connection,
    "excel.preview": handle_preview_excel,
    "import.run": handle_run_import,
}


def main() -> int:
    """Boucle principale : lit stdin ligne par ligne et dispatch."""
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue

        try:
            msg = json.loads(line)
        except json.JSONDecodeError as e:
            write_error("unknown", "INVALID_JSON", f"JSON invalide : {e}")
            continue

        request_id = msg.get("id", "unknown")
        method = msg.get("method")
        parameters = msg.get("params", {})

        if method not in HANDLERS:
            write_error(
                request_id,
                "UNKNOWN_METHOD",
                f"Méthode inconnue : {method}",
                f"Méthodes disponibles : {list(HANDLERS)}",
            )
            continue

        try:
            HANDLERS[method](request_id, parameters)
        except Exception as e:
            write_error(
                request_id,
                "INTERNAL_ERROR",
                str(e),
                traceback.format_exc(),
            )

    return 0


if __name__ == "__main__":
    sys.exit(main())
